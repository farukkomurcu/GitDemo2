# --HomePageData
import openpyxl

class HomePageData:

    #fixtures accept only list way even if have dict or tuple
    test_HomePage_data=[{"firstname":"Rahul","lastname":"shetty","gender":"male"},
                        {"firstname":"Anshika","lastname":"shetty","gender":"Female"}]

    @staticmethod #static method so it could removed self parameter from getTestData
                  #self is requires just only non statical method.
    def getTestData(test_case_name):
        Dict={}
        book = openpyxl.load_workbook("D:\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict] #it ll send in this squared bracket format